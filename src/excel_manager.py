"""
Модуль для работы с Excel файлами.
Содержит функции сохранения данных о точках в формате Excel с форматированием.
"""

import os
from typing import List

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from models.points import PointRecord


def save_points_without_city_to_csv(
    points: List[PointRecord], csv_path: str, log_message=None
) -> bool:
    """
    Сохраняет список точек без города в CSV-файл.

    Args:
        points (List[PointRecord]): Список точек без города
        csv_path (str): Путь к CSV-файлу
        log_message: Функция для логирования

    Returns:
        bool: True если успешно, False если ошибка
    """
    try:
        import csv
        import os

        # Ключ для уникальности: (date, time, latitude, longitude)
        def point_key(p):
            return (str(p.date), str(p.time), str(p.latitude), str(p.longitude))

        # Считываем уже существующие точки, если файл есть
        existing_keys = set()
        existing_rows = []
        if os.path.exists(csv_path):
            with open(csv_path, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    key = (
                        row.get("Date", ""),
                        row.get("Time", ""),
                        row.get("Lat_WGS84", ""),
                        row.get("Lon_WGS84", ""),
                        row.get("X_SK-42_Gauss_Kruger", ""),
                        row.get("Y_SK-42_Gauss_Kruger", ""),
                        row.get("Country_Value", ""),
                        row.get("City_Value", ""),
                        row.get("Description of the area", ""),
                        row.get("Description of the region", ""),
                        row.get("Original text", ""),
                        row.get("File_Path", ""),
                    )
                    existing_keys.add(key)
                    existing_rows.append(row)

        # Собираем новые уникальные точки
        new_rows = []
        for p in points:
            key = point_key(p)
            if key not in existing_keys:
                new_rows.append(
                    {
                        "Date": p.date,
                        "Time": p.time,
                        "Lat_WGS84": p.latitude,
                        "Lon_WGS84": p.longitude,
                        "X_SK-42_Gauss_Kruger": p.x_sk42,
                        "Y_SK-42_Gauss_Kruger": p.y_sk42,
                        "Country_Value": getattr(p, "country", ""),
                        "City_Value": getattr(p, "city", ""),
                        "Description of the area": getattr(p, "area_desc", ""),
                        "Description of the region": getattr(p, "region_desc", ""),
                        "Original text": getattr(p, "original_text", ""),
                        "File_Path": getattr(p, "file_path", ""),
                    }
                )
                existing_keys.add(key)

        header = [
            "Date",
            "Time",
            "Lat_WGS84",
            "Lon_WGS84",
            "X_SK-42_Gauss_Kruger",
            "Y_SK-42_Gauss_Kruger",
            "Country_Value",
            "City_Value",
            "Description of the area",
            "Description of the region",
            "Original text",
            "File_Path",
        ]

        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(header)
            # Сохраняем старые строки
            for row in existing_rows:
                writer.writerow(
                    [
                        row.get("Date", ""),
                        row.get("Time", ""),
                        row.get("Lat_WGS84", ""),
                        row.get("Lon_WGS84", ""),
                        row.get("X_SK-42_Gauss_Kruger", ""),
                        row.get("Y_SK-42_Gauss_Kruger", ""),
                        row.get("Country_Value", ""),
                        row.get("City_Value", ""),
                        row.get("Description of the area", ""),
                        row.get("Description of the region", ""),
                        row.get("Original text", ""),
                        row.get("File_Path", ""),
                    ]
                )
            # Сохраняем новые строки
            for row in new_rows:
                writer.writerow(
                    [
                        row.get("Date", ""),
                        row.get("Time", ""),
                        row.get("Lat_WGS84", ""),
                        row.get("Lon_WGS84", ""),
                        row.get("X_SK-42_Gauss_Kruger", ""),
                        row.get("Y_SK-42_Gauss_Kruger", ""),
                        row.get("Country_Value", ""),
                        row.get("City_Value", ""),
                        row.get("Description of the area", ""),
                        row.get("Description of the region", ""),
                        row.get("Original text", ""),
                        row.get("File_Path", ""),
                    ]
                )
        if log_message:
            log_message(
                f"Точки без города сохранены в {csv_path}",
                color="blue",
                logger_level="info",
            )
        return True
    except Exception as e:
        if log_message:
            log_message(
                f"Ошибка при сохранении точек без города в CSV: {e}",
                color="yellow",
                logger_level="warning",
            )
        return False


def save_points_to_excel(
    points_folder: List[PointRecord], data_xlsx_path: str, log_message=None
) -> bool:
    """
    Сохраняет список PointRecord в Excel с форматированием переносов строк.

    Args:
        points_folder (List[PointRecord]): Список объектов PointRecord для сохранения
        data_xlsx_path (str): Путь к файлу Excel для сохранения
        log_message: Функция для логирования

    Returns:
        bool: True если сохранение успешно, False если произошла ошибка
    """
    try:
        # Подготовка данных с нормализацией переносов строк и согласованными заголовками
        header = [
            "Date",
            "Time",
            "Lat_WGS84",
            "Lon_WGS84",
            "X_SK-42_Gauss_Kruger",
            "Y_SK-42_Gauss_Kruger",
            "Country_Value",
            "City_Value",
            "Description of the area",
            "Description of the region",
            "Original text",
            "File_Path",
        ]
        df_data = []
        for point in points_folder:
            row_data = {
                "Date": getattr(point, "date", ""),
                "Time": getattr(point, "time", ""),
                "Lat_WGS84": getattr(point, "latitude", ""),
                "Lon_WGS84": getattr(point, "longitude", ""),
                "X_SK-42_Gauss_Kruger": getattr(point, "x_sk42", ""),
                "Y_SK-42_Gauss_Kruger": getattr(point, "y_sk42", ""),
                "Country_Value": getattr(point, "country", ""),
                "City_Value": getattr(point, "city", ""),
                "Description of the area": getattr(point, "area_desc", ""),
                "Description of the region": getattr(point, "region_desc", ""),
                "Original text": str(getattr(point, "original_text", ""))
                .replace("\r\n", "\n")
                .replace("\r", "\n"),
                "File_Path": getattr(point, "file_path", ""),
            }
            df_data.append(row_data)
        points_df = pd.DataFrame(df_data, columns=header)

        # Сохранение с openpyxl для поддержки форматирования
        with pd.ExcelWriter(data_xlsx_path, engine="openpyxl") as writer:
            points_df.to_excel(writer, index=False, sheet_name="Points")
            worksheet = writer.sheets["Points"]

            # Стили для таблицы
            header_font = Font(bold=True, name="Times New Roman", size=11)
            data_font = Font(name="Times New Roman", size=10)
            header_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            center_alignment = Alignment(horizontal="center", vertical="center")
            wrap_alignment = Alignment(
                wrapText=True, vertical="center", horizontal="center"
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            # Форматирование заголовков (первая строка)
            for col_num in range(1, len(points_df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # Форматирование данных и границы для всех ячеек
            for row_num in range(2, worksheet.max_row + 1):
                for col_num in range(1, len(points_df.columns) + 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = thin_border
                    cell.alignment = center_alignment
                    cell.font = data_font

            # Специальное форматирование для столбца original_text
            original_text_col = None
            if "original_text" in points_df.columns:
                original_text_col = (
                    points_df.columns.get_indexer(["original_text"])[0] + 1
                )
                for row_num in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_num, column=original_text_col)
                    if cell.value:
                        cell.alignment = wrap_alignment
                        cell.font = data_font
                col_letter = worksheet.cell(
                    row=1, column=original_text_col
                ).column_letter
                worksheet.column_dimensions[col_letter].width = 50

            # Автоподбор ширины для всех остальных столбцов
            for col_num in range(1, len(points_df.columns) + 1):
                if original_text_col is None or col_num != original_text_col:
                    col_letter = worksheet.cell(row=1, column=col_num).column_letter
                    max_length = 0
                    for row in worksheet[col_letter]:
                        try:
                            if len(str(row.value)) > max_length:
                                max_length = len(str(row.value))
                        except Exception:
                            pass
                    adjusted_width = min(max(max_length + 2, 8), 50)
                    worksheet.column_dimensions[col_letter].width = adjusted_width

        if log_message:
            log_message(
                f"Excel файл успешно сохранён: {data_xlsx_path}",
                level="info",
                color="blue",
            )
        return True
    except Exception as e:
        if log_message:
            log_message(
                f"Ошибка при сохранении Excel: {e}", level="warning", color="yellow"
            )
        return False
